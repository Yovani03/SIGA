import openpyxl

def excel_to_html(wb):
    ws = wb.active
    html = '<!DOCTYPE html>\n<html>\n<head>\n<meta charset="UTF-8">\n<style>\n'
    html += '    @page {\n        size: letter landscape;\n        margin: 1cm;\n    }\n'
    html += '    body {\n        font-family: sans-serif;\n    }\n'
    html += '</style>\n</head>\n<body>\n'
    html += '<table style="border-collapse: collapse; width: 100%; font-family: sans-serif; font-size: 10px; text-align: center;">\n'
    
    merged_cells = ws.merged_cells.ranges
    skip_cells = set()
    
    for row in ws.iter_rows():
        html += '  <tr>\n'
        for cell in row:
            if cell.coordinate in skip_cells:
                continue
                
            colspan = 1
            rowspan = 1
            for merged in merged_cells:
                if cell.coordinate == merged.start_cell.coordinate:
                    colspan = merged.max_col - merged.min_col + 1
                    rowspan = merged.max_row - merged.min_row + 1
                    
                    for r in range(merged.min_row, merged.max_row + 1):
                        for c in range(merged.min_col, merged.max_col + 1):
                            if not (r == merged.min_row and c == merged.min_col):
                                skip_cells.add(ws.cell(row=r, column=c).coordinate)
                    break
                    
            val = cell.value if cell.value is not None else ''
            # Handle newlines in val
            val = str(val).replace('\n', '<br>')
            
            style = []
            
            # Background Color
            if cell.fill and cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb') and type(cell.fill.start_color.rgb) == str:
                rgb = cell.fill.start_color.rgb
                if rgb != '00000000':
                    style.append(f'background-color: #{rgb[2:]}')
                    
            # Font
            if cell.font:
                if cell.font.bold: style.append('font-weight: bold')
                if cell.font.size: style.append(f'font-size: {int(cell.font.size)}px')
                if cell.font.name: style.append(f'font-family: "{cell.font.name}", sans-serif')
            
            # Alignment
            if cell.alignment:
                if cell.alignment.horizontal: style.append(f'text-align: {cell.alignment.horizontal}')
                if cell.alignment.vertical:
                    v_align = cell.alignment.vertical
                    if v_align == 'center': v_align = 'middle'
                    style.append(f'vertical-align: {v_align}')
            else:
                style.append('text-align: center')
                style.append('vertical-align: middle')
                
            # Borders
            b_style = '1px solid black'
            if cell.border:
                if cell.border.top and cell.border.top.style: style.append(f'border-top: {b_style}')
                if cell.border.bottom and cell.border.bottom.style: style.append(f'border-bottom: {b_style}')
                if cell.border.left and cell.border.left.style: style.append(f'border-left: {b_style}')
                if cell.border.right and cell.border.right.style: style.append(f'border-right: {b_style}')
            else:
                # Default border if none specified just so it looks like a grid
                style.append(f'border: {b_style}')
                
            style.append('padding: 2px')
            
            style_str = '; '.join(style) + ';'
            html += f'    <td colspan="{colspan}" rowspan="{rowspan}" style="{style_str}">{val}</td>\n'
            
        html += '  </tr>\n'
    html += '</table>\n</body>\n</html>'
    return html

wb = openpyxl.load_workbook(r'C:\Autrotransportes\backend\logistica\plantillas\FORMATO CEDIS Y TRANSPORTES FINAL.xlsx')
html = excel_to_html(wb)
with open('c:/Autrotransportes/backend/logistica/templates/bitacora_base.html', 'w', encoding='utf-8') as f:
    f.write(html)
print('HTML generated')
